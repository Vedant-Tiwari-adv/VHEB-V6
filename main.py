import tkinter as tk
import customtkinter,math
from openpyxl import load_workbook
from PIL import Image,ImageTk
from Google import Create_Service
import os, io
from googleapiclient.http import MediaIoBaseDownload, MediaFileUpload#pip install google-api-python-client
import time,pyautogui,datetime#pip install google-auth-oauthlib
def datestr():
    today=datetime.datetime.now()
    months=['January', 'February'
            , 'March', 'April',
            'May', 'June',
            'July','August',
            'September','October',
            'November', 'December']
    return str(months[today.month-1])+' '+str(today.year)
def get_lst_xl_vals():
    wb_obj=load_workbook(filename='Electricity.xlsx', data_only=True)
    sheet_obj=wb_obj.active
    lst_readings=[]
    
    for column in range(sheet_obj.max_column, 100, -1):
        if column%4==1:
            val=sheet_obj.cell(row=3,column=column).value
            if val is not None:
                column-=2
                break
    NagarNigamReading=sheet_obj.cell(row=3,column=column+2).value
    for rw in [6,7,8,9,10,11,12]:
        if rw == 7:
            lst_readings.append(None)
        else:
            lst_readings.append(int(sheet_obj.cell(row=rw,column=column).value))
    return lst_readings, NagarNigamReading
def update_xlsheet(lst_readings, crnt_readings,lst_NagarNigamReading,crnt_NagarNigamReading,NagarNigamBillAmt, unitprice=7.2):
    '''lst_readings=[c6 val,none,c8.c9,c10,c11,c12] from get_lst_xl_vals
        crnt_reading=[mcb,none,,rightmost,no1,no2,no3,motor] neeed from ui
        unitprice need form ui
        crnt_nagarnigamreading need form ui

        lstred,finalred=get_lst_xl_vals()
        update_xlsheet(lstred,[8500,None, 1400,2600,2100,1600,200],finalred,30000,7.3,70000)
    '''
    print("Updating Values.")
    #finding column to edit
    wb_obj=load_workbook(filename='Electricity.xlsx', data_only=True)
    sheet_obj=wb_obj.active
    for column in range(sheet_obj.max_column, 100, -1):
        if column%4==1:
            val=sheet_obj.cell(row=3,column=column).value
            if val is not None:
                column+=2
                break
    #editing the values
    sheet_obj.cell(row=1,column=column).value=datestr()
    print('Column Found : ',column)
    t_unit=0
    t_price=0
    row=[6,7,8,9,10,11,12]
    for i in row:
        if i==7:
            continue
        else:
            unit=crnt_readings[row.index(i)]-lst_readings[row.index(i)]
            sheet_obj.cell(row=i,column=column+1).value=unit
            t_unit+=unit
            sheet_obj.cell(row=i,column=column).value=crnt_readings[row.index(i)]
            sheet_obj.cell(row=i,column=column+2).value=round(unit*unitprice,1)
            t_price+=unit*unitprice

    sheet_obj.cell(row=7,column=column+1).value=crnt_NagarNigamReading-lst_NagarNigamReading-t_unit
    sheet_obj.cell(row=7,column=column+2).value=round((crnt_NagarNigamReading-lst_NagarNigamReading-t_unit)*unitprice,1)
    t_price+=(crnt_NagarNigamReading-lst_NagarNigamReading-t_unit)*unitprice
    sheet_obj.cell(row=2,column=column+2).value=unitprice
    sheet_obj.cell(row=3,column=column+2).value=crnt_NagarNigamReading
    sheet_obj.cell(row=4,column=column+2).value=NagarNigamBillAmt
    sheet_obj.cell(row=13,column=column+2).value=round(t_price,1)
    print("Values Added to excell file.")
    wb_obj.save(filename="Electricity.xlsx")
    wb_obj.close()    
File_Id='1nGcHJPiHYEjmnfkwf5Z8JA7GNu1WSTVP'
def file_download():
    CLIENT_SECRET_FILE='credentials.json'
    API_NAME='drive'
    API_VERSION='v3'
    SCOPES=['https://www.googleapis.com/auth/drive']

    service=Create_Service(CLIENT_SECRET_FILE, API_NAME, API_VERSION, SCOPES)

    FILE_ID=File_Id
    FILE_NAME='Electricity.xlsx'

    request=service.files().get_media(fileId=FILE_ID)
    fh=io.BytesIO()
    downloader= MediaIoBaseDownload(fd=fh,request=request)
    done=False
    #itterating through chunks of file and downloading them
    #the last  chunk will set "done" as True
    while not done:
        status, done=downloader.next_chunk()
        print('Download Complete' )
    fh.seek(0)
    #this saves the file
    #if file allready exists it ovver writes !!!!!!!
    with open(FILE_NAME, 'wb') as f:
        f.write(fh.read())
        f.close()
def file_upload():
    print("Uploading file")
    CLIENT_SECRET_FILE='credentials.json'
    API_NAME='drive'
    API_VERSION='v3'
    SCOPES=['https://www.googleapis.com/auth/drive']
    service=Create_Service(CLIENT_SECRET_FILE, API_NAME, API_VERSION, SCOPES)
    file_id=File_Id#id of file u want to replace
#electricity1 is the new file, that has to be put in place of electricity,pdf
    media_content= MediaFileUpload('Electricity.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    service.files().update(
        fileId=file_id,
        media_body= media_content).execute()
    print("Upload Complete.")
def changeWLine():
    pyautogui.keyDown('shift')
    pyautogui.press("enter")
    pyautogui.keyUp('shift')
def whatsapp_processing(Units,Amounts,unitprice=7.2,waterbool=True,tenants=6):
    '''whatsapp_processing([371,247,37,39,64,39,39],[2708.3,1803.1,270.1,284.7,467.2,284.7,284.7],7.3)
    amount=[c6toc13] same with units
    '''
    print("Pyautogui Active \nOPEN WHATSAPP")
    for sec in range(10,0,-1):
        time.sleep(1)
        print("Typing bills in : ",sec, "seconds")
    print("pyautogui Active")
    flat=['Front Ground','Back Ground',
          'Balcony Room','Hall Upper',
          'Middle Upper','Back Upper']
    if waterbool:
        wamt=round(Amounts[-1]/tenants)
    for i in range(len(flat)):
        pyautogui.write("*Electricity Bill for {}*".format(flat[i]))
        changeWLine()
        pyautogui.write("{}    |    _Rs{}/Unit_".format(datestr(),unitprice))
        changeWLine()
        pyautogui.write('Units            :     {}'.format(Units[i]))
        changeWLine()
        pyautogui.write('Amount        :     Rs{}'.format(round(Amounts[i],1)))
        changeWLine()
        if waterbool:
            pyautogui.write('Water           :     Rs{}'.format(wamt))

            changeWLine()
            pyautogui.write('*'+'Total            :     Rs{}*'.format(round(float(Amounts[i])+float(wamt))))
            pyautogui.press("enter")
        else:
            pyautogui.press("enter")
def read_xlsheet(seclast=False):
    '''seclast is bool if true then we geting last only'''
    wb_obj=load_workbook(filename='Electricity.xlsx', data_only=True)
    sheet_obj=wb_obj.active
    for column in range(sheet_obj.max_column, 100, -1):
        if column%4==1:
            val=sheet_obj.cell(row=3,column=column).value
            if val is not None:
                column-=2
                break
    
    if seclast:
        column-=4
    Units=[]
    Amounts=[]
    for row in [6,7,8,9,10,11,12]:
        Units.append(int(sheet_obj.cell(row=row,column=column+1).value))
        Amounts.append(int(sheet_obj.cell(row=row,column=column+2).value))
    return Units,Amounts,float(sheet_obj.cell(row=2,column=column+2).value)
def suggested_unitprice(margin=6):
    #margin is 8 bydefault can adjust
    wb_obj=load_workbook(filename='Electricity.xlsx', data_only=True)
    sheet_obj=wb_obj.active
    for column in range(sheet_obj.max_column, 100, -1):
        if column%4==1:
            val=sheet_obj.cell(row=3,column=column).value
            if val is not None:
                column-=2
                break
    Units=[]
    for row in [6,7,8,9,10,11,12]:
        Units.append(int(sheet_obj.cell(row=row,column=column+1).value))
    unit=sum(Units)
    bill_amt=int(sheet_obj.cell(row=4,column=column+2).value)
    price=(bill_amt+(bill_amt*(margin/100)))/unit
    return "Rs "+str(round(price,2))
def delete_data():
        wb_obj=load_workbook(filename='Electricity.xlsx', data_only=True)
        sheet_obj=wb_obj.active
        for column in range(sheet_obj.max_column, 100, -1):
            if column%4==1:
                val=sheet_obj.cell(row=3,column=column).value
                if val is not None:
                    column-=2
                    break
        cl=[6,8,9,10,11,12]
        cl1=[6,7,8,9,10,11,12]
        cl2=[2,3,4,6,7,8,9,10,11,12,13]
        for row in cl:
                sheet_obj.cell(row=row,column=column).value=None
        for row in cl1:
            sheet_obj.cell(row=row,column=column+1).value=None
        for row in cl2:
            sheet_obj.cell(row=row,column=column+2).value=None
        wb_obj.save(filename="Electricity.xlsx")
        wb_obj.close() 


def __main__():
    #file_download()
    global update_Frame
    bill_bool=0
    Intro_bool=1
    def wateru_need_fn():
        if wateru_box.get():
            
            global wateru_field, water_lable
            wateru_field=customtkinter.CTkEntry(checku_Frame, placeholder_text="No. of Tenants",placeholder_text_color='grey',height=15,width=100,corner_radius=5)
            wateru_field.grid(row=2, column=1, padx=20, pady=5)
            water_lable=customtkinter.CTkLabel(checku_Frame,text='6 by default',text_color='grey',font=('Verdana',10))
            water_lable.grid(row=3,column=1,padx=10,pady=5 )
        if not wateru_box.get():
            try:
                water_lable.destroy()
                wateru_field.destroy()
            except:
                pass
    def water_need_fn():
        if water_box.get():
            
            global water_field, water_lable
            water_field=customtkinter.CTkEntry(check_Frame, placeholder_text="No. of Tenants",placeholder_text_color='grey',height=15,width=100,corner_radius=5)
            water_field.grid(row=2, column=1, padx=20, pady=5)
            water_lable=customtkinter.CTkLabel(check_Frame,text='6 by default',text_color='grey',font=('Verdana',10))
            water_lable.grid(row=3,column=1,padx=10,pady=5 )
        if not water_box.get():
            try:
                water_lable.destroy()
                water_field.destroy()
            except:
                pass
    def unitpriceu_box_fn():
        if unitpriceu_box.get():
            global unitpriceu_field ,suggest_lable
            unitpriceu_field=customtkinter.CTkEntry(checku_Frame, placeholder_text="Enter price per unit",placeholder_text_color='grey',height=15,width=150,corner_radius=5)
            unitpriceu_field.grid(row=5, column=1, padx=20, pady=5)
            txt='Suggested Unit Price : {}'.format(suggested_unitprice())
            suggest_lable=customtkinter.CTkLabel(checku_Frame,text=txt,text_color='grey',font=('Verdana',10))
            suggest_lable.grid(row=6,column=1,padx=10,pady=5 )
        if not unitpriceu_box.get():
            try:
                unitpriceu_field.destroy()
                suggest_lable.destroy()
            except:
                pass
    def unitprice_box_fn():
        if unitprice_box.get():
            global unitprice_field ,suggest_lable
            unitprice_field=customtkinter.CTkEntry(check_Frame, placeholder_text="Enter price per unit",placeholder_text_color='grey',height=15,width=150,corner_radius=5)
            unitprice_field.grid(row=5, column=1, padx=20, pady=5)
            txt='Suggested Unit Price : {}'.format(suggested_unitprice())
            suggest_lable=customtkinter.CTkLabel(check_Frame,text=txt,text_color='grey',font=('Verdana',10))
            suggest_lable.grid(row=6,column=1,padx=10,pady=5 )
        if not unitprice_box.get():
            try:
                unitprice_field.destroy()
                suggest_lable.destroy()
            except:
                pass
    def print_event_inbill():
        try:
            noti_Frame.destroy()

        except:
            pass
        units,amts,l=read_xlsheet()
        if unitprice_box.get() and water_box.get():
            whatsapp_processing(units,amts,unitprice_int,waterbool=wbool,tenants=nten)
        elif unitprice_box.get()==1 and water_box.get()==0:
            whatsapp_processing(units,amts,unitprice_int,waterbool=wbool)
        elif unitprice_box.get()==0 and water_box.get()==1:
            whatsapp_processing(Units=units,Amounts=amts,waterbool=wbool,tenants=nten)
        elif unitprice_box.get()==0 and water_box.get()==0:
            whatsapp_processing(Units=units,Amounts=amts,waterbool=wbool)
        global noti_body_text
        noti_body_text='All the bills have been\nprinted successfully.'
        #noti_dummy.toggle()
        bill_button.configure(state='normal')
        print_button.configure(state='normal')
        delete_button.configure(state='normal')
        upload_button.configure(state='normal')
        intro_button.configure(state='normal')
        intro_bool_dummy.toggle()
        

    def submit_button_event():
        global popup_body_text,popup_title_text

        popup_title_text='              Upload Data Confirmation'
        popup_body_text='     Please make sure--                     \n1) All values entered, are numeric.\n2) Tick  boxes for water bill and    \n       Unitprice are appropriately filledup.'
        try:
            final_upload_button.configure(width=0)
        except:
            pass
        try:
            idelete_button.configure(width=0)
        except:
            pass
        
        popup_bool_dummy.toggle()





    
    def final_uploadb_event():
        global popup_body_text,popup_title_text
        popup_title_text='              Upload Data Confirmation'
        popup_body_text='     Please make sure--                     \n1) All values entered, are numeric.\n2) Tick  boxes for water bill and    \n       Unitprice are appropriately filledup.'
        try:
            submit_button.configure(width=0)
        except:
            pass
        try:
            idelete_button.configure(width=0)
        except:
            pass
        popup_bool_dummy.toggle()

    def disable_options(rev):

        if not rev:
            intro_button.configure(state='disabled')

            bill_button.configure(state='disabled')

            print_button.configure(state='disabled')

            upload_button.configure(state='disabled')

            delete_button.configure(state='disabled')

        else:

            intro_button.configure(state='normal')

            bill_button.configure(state='normal')

            print_button.configure(state='normal')

            upload_button.configure(state='normal')

            delete_button.configure(state='normal')

        





        
    print("avi")
    
    app=customtkinter.CTk()
    app.title("  VHEB v3")
    app.geometry("600x500")
    app.resizable(False,False)
    app.configure(bg='000000')
    customtkinter.set_appearance_mode("dark")
    customtkinter.set_default_color_theme("dark-blue")
    def uupload_button_event():
        try:
            noti_Frame.destroy()
            
        except:
            pass
        upload_button.configure(state='disabled',fg_color="#141414")
        delete_button.configure(state='normal', fg_color='transparent')
        print_button.configure(state='normal', fg_color='transparent')
        intro_button.configure(state='normal', fg_color='transparent')
        bill_button.configure(state='normal', fg_color='transparent')
        try:
            Main_Frame.cget("width")
            Main_Frame.destroy()
        except:
            pass

        try:
            Intro_Frame.cget('width')
            Intro_Frame.destroy()
        except:
            pass
        try:
            Delete_Frame.cget('width')
            Delete_Frame.destroy()
        except:
            pass
        try:
            noti_Frame.destroy()
        except:
            pass
        update_dummy.toggle()
        #btitle.destroy()
        #update_logo.destroy()
        #submit_button.configure(width=0)
        #submit_button.destroy()
        #print('buttonn destroyed: ',submit_button.cget('width'))



#########################################################




    def update_page():
        try:
            global update_Frame
            update_Frame.destroy()
        except:
            pass
        try:
            Print_Frame.destroy()
        except:
            pass
        try:
            Intro_Frame.cget('width')
            Intro_Frame.destroy()
        except:
            pass
        try:
            Delete_Frame.cget('width')
            Delete_Frame.destroy()
        except:
            pass
        global entry1u,entry2u,entry3u,entry4u,entry5u,entry6u,entry7u,unitpriceu_box,wateru_box,entry8u,final_upload_button, checku_Frame,entryu_Frame
        update_Frame=customtkinter.CTkFrame(master=app,corner_radius=10 ,fg_color='#000000',width=470,height=490)

        update_Frame.place(x=120,y=10)                    
        entryu_Frame=customtkinter.CTkFrame(master=update_Frame, fg_color='#141414',corner_radius=10)    
        checku_Frame=customtkinter.CTkFrame(master=update_Frame, fg_color='#141414',corner_radius=10)        
        entryu_Frame.grid_columnconfigure((0), weight=1)
        checku_Frame.grid_columnconfigure((0), weight=1)
        entryu_Frame.pack(padx=(10,5),pady=(120,10) , side='left',fill='both')
        checku_Frame.pack(padx=(5,10),pady=(120,130) )

        Uploadimgb =ImageTk.PhotoImage(Image.open("assets\\upload.png"))
        upload_logo=customtkinter.CTkButton(update_Frame,state='disabled',text='',image=Uploadimgb, border_width=0,fg_color='transparent')
        upload_logo.place(x=55,y=15)
        utitle = customtkinter.CTkLabel(update_Frame,text='Enter data to be uploaded\nto xlsx file, stored in gdrive.',text_color='grey',font=('Verdana',18))
        utitle.place(x=160,y=30)
        uploadbimg=ImageTk.PhotoImage(Image.open("assets\\uploadb.png").resize((40,40)))

        final_upload_button=customtkinter.CTkButton(update_Frame,image= uploadbimg,
                                        corner_radius=15,border_spacing=1,text='Upload',
                                        fg_color='#008AA9',command=final_uploadb_event,
                                        height=68,width=100,  border_width=0,
                                        compound='left',font=('Verdana',14)) 
        final_upload_button.place(x=300,y=380 )
    
        btitleu_entry=customtkinter.CTkLabel(entryu_Frame,text='Enter current readings for :',text_color='grey',font=('Verdana',14))
        btitleu_entry.grid(row=0, column=1, padx=20, pady=(15,5), sticky='nw')

        entry1u=customtkinter.CTkEntry(entryu_Frame, placeholder_text="Meter No. 1",placeholder_text_color='grey',height=15,width=150,corner_radius=5)
        entry1u.grid(row=1, column=1, padx=20, pady=(0,5))

        entry2u=customtkinter.CTkEntry(entryu_Frame, placeholder_text="Meter No. 2",placeholder_text_color='grey',height=15,width=150,corner_radius=5)
        entry2u.grid(row=2, column=1, padx=20, pady=5)

        entry3u=customtkinter.CTkEntry(entryu_Frame, placeholder_text="Meter No. 3",placeholder_text_color='grey',height=15,width=150,corner_radius=5)
        entry3u.grid(row=3, column=1, padx=20, pady=5)

        entry4u=customtkinter.CTkEntry(entryu_Frame, placeholder_text="Meter No. 4",placeholder_text_color='grey',height=15,width=150,corner_radius=5)
        entry4u.grid(row=4, column=1, padx=20, pady=5)

        entry5u=customtkinter.CTkEntry(entryu_Frame, placeholder_text="Meter No. 5",placeholder_text_color='grey',height=15,width=150,corner_radius=5)
        entry5u.grid(row=5, column=1, padx=20, pady=5)

        entry6u=customtkinter.CTkEntry(entryu_Frame, placeholder_text="Meter No. 0",placeholder_text_color='grey',height=15,width=150,corner_radius=5)
        entry6u.grid(row=6, column=1, padx=20, pady=5)

        title_entry1=customtkinter.CTkLabel(entryu_Frame,text='Values form NagarNigam bill :',text_color='grey',font=('Verdana',14))
        title_entry1.grid(row=7, column=1, padx=20, pady=(15,5))
        
        entry7u=customtkinter.CTkEntry(entryu_Frame, placeholder_text="Current Reading",placeholder_text_color='grey',height=15,width=150,corner_radius=5)
        entry7u.grid(row=8, column=1, padx=20, pady=5)

        entry8u=customtkinter.CTkEntry(entryu_Frame, placeholder_text="Total Bill Amount",placeholder_text_color='grey',height=15,width=150,corner_radius=5)
        entry8u.grid(row=9, column=1, padx=20, pady=(5,10))
        

        wateru_box=customtkinter.CTkCheckBox(checku_Frame,height=2,fg_color='#37D0EE',width=2,text='Charge for water',text_color='grey',font=('Verdana',14)
                                             ,command=wateru_need_fn)
        wateru_box.grid(row=1,column=1,padx=20,pady=10)
        wateru_box.toggle()
        unitpriceu_box=customtkinter.CTkCheckBox(checku_Frame,height=2,fg_color='#37D0EE',width=2,text='Change Unit Price',text_color='grey',font=('Verdana',14)
                                                 ,command=unitpriceu_box_fn)
        unitpriceu_box.grid(row=4,column=1,padx=20,pady=10)
        unitpriceu_box.toggle()

#########################################################
        
        
        
    def bill_button_event():
        try:
            noti_Frame.destroy()
        except:
            pass
        bill_button.configure(state='disabled',fg_color="#141414")
        delete_button.configure(state='normal', fg_color='transparent')
        print_button.configure(state='normal', fg_color='transparent')
        intro_button.configure(state='normal', fg_color='transparent')
        upload_button.configure(state='normal', fg_color='transparent')
        try:
            Intro_Frame.destroy()         
        except:
            print("Element allready deleted")


        bill_bool_dummy.toggle()
        
    def prnt_button_event():
        try:
            noti_Frame.destroy()
        except:
            pass
        print_button.configure(state='disabled',fg_color="#141414")
        bill_button.configure(state='normal', fg_color='transparent')
        delete_button.configure(state='normal', fg_color='transparent')
        intro_button.configure(state='normal', fg_color='transparent')
        upload_button.configure(state='normal', fg_color='transparent')
        try:
            Main_Frame.cget("width")
            Main_Frame.destroy()
        except:
            pass
        try:
            update_Frame.destroy()
        except:
            pass
        try:
            Intro_Frame.cget('width')
            Intro_Frame.destroy()
        except:
            pass
        try:
            Delete_Frame.cget('width')
            Delete_Frame.destroy()
        except:
            pass
        print_button_event()
        print_bool_dummy.toggle()

        
        
        pass
    def print_button_event():
        def print_event_inprint():
            u,a,up=read_xlsheet(lst_box.get())
            whatsapp_processing(u, a, up, water_box_inbill.get())
            intro_bool_dummy.toggle()
        global Print_Frame, print_bool_dummy
        Print_Frame=customtkinter.CTkFrame(master=app,corner_radius=10,fg_color='#000000',width=470,height=480)
        print_bool_dummy=customtkinter.CTkCheckBox(Print_Frame, command=print_button_event)


        Print_Frame.place(x=120,y=10)
        Printimgb =ImageTk.PhotoImage(Image.open("assets\\Print.png").resize((64,64)))
        print_logo=customtkinter.CTkButton(Print_Frame,text='',state='disabled',image=Printimgb, border_width=0,fg_color='transparent')
        print_logo.place(x=25,y=5)
        ptitle = customtkinter.CTkLabel(Print_Frame,text='Generate Bill for already\nuploaded data, on Whatsapp.',text_color='grey',font=('Verdana',18))
        ptitle.place(x=140,y=16)


        hint1_img =ImageTk.PhotoImage(Image.open("assets\\hint1.png").resize((250,120)))
        hint2_img= ImageTk.PhotoImage(Image.open("assets\\hint2.png").resize((250,120)))
        hint1_logo=customtkinter.CTkButton(Print_Frame,image=hint1_img,text='' ,border_width=0,fg_color='transparent')
        hint1_logo.place(x=7,y=80)
        hint1_logo.configure(state='disabled')
        hint2_logo=customtkinter.CTkButton(Print_Frame,image=hint2_img,text='', border_width=0,fg_color='transparent')
        hint2_logo.place(x=7,y=279)
        hint2_logo.configure(state='disabled')
        whatsapp_print_button= customtkinter.CTkButton(Print_Frame,image=Printimg,
                                                            text='Print', border_width=0,text_color='grey',
                                                            fg_color='#141414',corner_radius=15,border_spacing=1,
                                                                command=print_event_inprint,
                                                                    height=68,width=80,  hover_color='#424242',
                                                                compound='left',font=('Verdana',12))
        whatsapp_print_button.place(x=345,y=400)       
        hint1_Bill=customtkinter.CTkLabel(Print_Frame,text='Open  Whatsapp  web\nand click on a contact,\n as  shown  in  img1.\n*Note*\nAll the  bills will be sent\nto the selected contact.',text_color='grey',font=('Verdana',14))
        hint3_Bill=customtkinter.CTkLabel(Print_Frame,text='Be patient, it takes about 30s to type\nup all the bills.  This  process  will\nbe automated  in future iterations.',text_color='grey',font=('Verdana',13))
        hint4_Bill=customtkinter.CTkLabel(Print_Frame,text='Step 1',text_color='white',font=('Verdana',15))
        hint5_Bill=customtkinter.CTkLabel(Print_Frame,text='Step 3',text_color='white',font=('Verdana',15))
        hint6_Bill=customtkinter.CTkLabel(Print_Frame,text='Step 2',text_color='white',font=('Verdana',15))
        hint7_bill=customtkinter.CTkLabel(Print_Frame,text='Tick the following boxes,\nto the right as needed.',text_color='grey',font=('Verdana',14))
        hint2_Bill=customtkinter.CTkLabel(Print_Frame,text=' Click on Print button\nbelow, quickly change\nto   whatsapp   and\nclick on  text  box as\nshown   in   img2. ',text_color='grey',font=('Verdana',14))
        hint1_Bill.place(x=290,y=95)
        hint2_Bill.place(x=290,y=305)
        hint3_Bill.place(x=40,y=410)
        hint5_Bill.place(x=340,y=280)
        hint4_Bill.place(x=340,y=68)
        hint6_Bill.place(x=70,y=210)
        hint7_bill.place(x=15,y=235)
        water_box_inbill=customtkinter.CTkCheckBox(Print_Frame,height=2,fg_color='#37D0EE',width=2,text='Charge the tenants for the water.',text_color='grey',font=('Verdana',14))
        lst_box=customtkinter.CTkCheckBox(Print_Frame,height=2,fg_color='#37D0EE',width=2,text='Print second to last months bills.',text_color='grey',font=('Verdana',14))
        lst_box.place(x=200,y=220)
        water_box_inbill.place(x=200,y=250)
    



        
    def del_button_event():
        try:
            noti_Frame.destroy()
        except:
            pass
        try:
            Main_Frame.cget("width")
            Main_Frame.destroy()
        except:
            pass
        try:
            Intro_Frame.cget('width')
            Intro_Frame.destroy()
        except:
            pass
        try:
            Print_Frame.destroy()
        except:
            pass
        try:
            noti_Frame.destroy()
        except:
            pass
        bill_button.configure(state='normal', fg_color='transparent')
        print_button.configure(state='normal', fg_color='transparent')
        intro_button.configure(state='normal', fg_color='transparent')
        upload_button.configure(state='normal', fg_color='transparent')
        delete_button.configure(state='disabled',fg_color="#141414")
        delete_dummy.toggle()

        #popup_bool_dummy.toggle()

        
    def delete_page():
        global Delete_Frame
        Delete_Frame=customtkinter.CTkFrame(master=app,corner_radius=10,fg_color='#000000',width=470,height=480)
        Delete_Frame.place(x=120,y=10)
        def del_event_indel():
            global popup_body_text, popup_title_text
            popup_title_text='                     Are You Sure?'
            popup_body_text='\n       This will delete the last\n        data entry in the database.'

            try:
                submit_button.configure(width=0)
            except:
                pass
            try:
                final_upload_button.configure(width=0)
            except:
                pass
            popup_bool_dummy.toggle()

        dtitle = customtkinter.CTkLabel(Delete_Frame,text='Delete Last Entry\nData set.',text_color='grey',font=('Verdana',18))
        dtitle.place(x=200,y=16)
        deleteimgb=ImageTk.PhotoImage(Image.open("assets\\delete.png").resize((64,64)))
        deleteimgs=ImageTk.PhotoImage(Image.open("assets\\delete.png").resize((40,40)))
        delete_logo=customtkinter.CTkButton(Delete_Frame,text='',state='disabled',image=deleteimgb, border_width=0,fg_color='transparent')
        delete_logo.place(x=45,y=5)
        global idelete_button
        idelete_button= customtkinter.CTkButton(Delete_Frame,image=deleteimgs,
                                                        text='Delete', border_width=0,text_color='grey',
                                                        fg_color='#141414',corner_radius=15,border_spacing=1,
                                                        command=del_event_indel,
                                                        height=68,width=80,  hover_color='#424242',
                                                        compound='left',font=('Verdana',12))
        
        idelete_button.place(x=200,y=200)


        
    
    





    

    Option_Frame=customtkinter.CTkFrame(master=app,corner_radius=10,fg_color='#000000',height=482,width=100)
    Option_Frame.grid_columnconfigure((0), weight=1)
    #Option_Frame.pack(padx=(10,10),pady=10 , side='left', fill='y')
    Option_Frame.place(x=10,y=10)
    delete_dummy=customtkinter.CTkCheckBox(Option_Frame,command=delete_page)

    def noti_anim_in():

        global IN

        IN+=1
        angles=math.sin(math.radians(IN))
        noti_Frame.place(x=610+370*angles,y=20)
        if angles!=-1:
            #xcord-=abs(angles)
            app.after(4,noti_anim_in)
    def noti_anim_out():

        global OUT
        OUT+=1
        
        angles=math.sin(math.radians(OUT))
        r=610-370*angles
        noti_Frame.place(x=r,y=20)
        

        if int(r)==380:
            noti_Frame.destroy()
        if angles!=-1:
            #xcord-=abs(angles)
            app.after(4,noti_anim_out)
        
    def gnoti_anim_in():

        global INg

        INg+=1
        angles=math.sin(math.radians(IN))
        drive_Frame.place(x=610+370*angles,y=90)
        if angles!=-1:
            #xcord-=abs(angles)
            app.after(4,gnoti_anim_in)
    def gnoti_anim_out():

        global OUTg
        OUTg+=1
        
        angles=math.sin(math.radians(OUTg))
        r=610-370*angles
        drive_Frame.place(x=r,y=90)
        

        if int(r)==380:
            drive_Frame.destroy()
        if angles!=-1:
            #xcord-=abs(angles)
            app.after(4,gnoti_anim_out)

    def drive_status():
        def drive_noti_close_event():
            try:
                gnoti_anim_out()
            except:
                pass
        global OUTg,INg,drive_body_test,drive_Frame
        OUTg=90
        INg=180
        drive_Frame=customtkinter.CTkFrame(master=app,corner_radius=10,bg_color='#000000', border_width=2,border_color="white",
                                           fg_color='#37D0EE',height=60,width=340)
        drive_body=customtkinter.CTkLabel(drive_Frame,text=drive_body_text,
                                text_color='black',font=('Verdana',14))
        noimg=ImageTk.PhotoImage(Image.open("assets\\no1.png").resize((30,30)))
        driveimg=ImageTk.PhotoImage(Image.open("assets\\drive.png").resize((45,45)))
        no_button=customtkinter.CTkButton(drive_Frame,image= noimg,
                                        corner_radius=7,border_spacing=0,text=None,hover_color='red',
                                        fg_color='transparent',command=drive_noti_close_event,border_color="#000000",
                                        height=45,width=10,  border_width=2) 

        drive=customtkinter.CTkButton(drive_Frame,image= driveimg,
                                        corner_radius=7,border_spacing=0,text=None,
                                        fg_color='transparent',state='disabled',
                                         height=45,width=10, border_width=0)
        drive.place(x=3,y=3)
        no_button.place(x=286,y=8 )    
        drive_body.place(x=50,y=12)
        
        gnoti_anim_in()

    def notification():
        def noti_close_event():
            try:
                noti_anim_out()
                #noti_Frame.destroy()
            except:
                pass
    
        global noti_Frame, noti_body_text,IN,OUT
        OUT=90
        IN=180

        try:
            noti_Frame.destroy()
        except:
            pass
        noti_Frame=customtkinter.CTkFrame(master=app,corner_radius=10,bg_color='#000000', border_width=2,border_color="white",
                                           fg_color='#37D0EE',height=60,width=340)
        noti_body=customtkinter.CTkLabel(noti_Frame,text=noti_body_text,
                                text_color='black',font=('Verdana',14))
        noimg=ImageTk.PhotoImage(Image.open("assets\\no1.png").resize((30,30)))
        no_button=customtkinter.CTkButton(noti_Frame,image= noimg,
                                        corner_radius=7,border_spacing=0,text=None,hover_color='red',
                                        fg_color='transparent',command=noti_close_event,border_color="#000000",
                                        height=45,width=10,  border_width=2) 

        no_button.place(x=286,y=8 )    
        noti_body.place(x=30,y=12)


        noti_anim_in()
    drive_dummy=customtkinter.CTkCheckBox(Option_Frame,command=drive_status)
    global noti_dummy
    noti_dummy=customtkinter.CTkCheckBox(Option_Frame,command=notification)
    def popup():
        disable_options(False)
        try:
            submit_button.configure(state="disabled")
        except:
            pass
        try:
            final_upload_button.configure(state="disabled")
        except:
            pass
        
        #disabel other buttons when popup comes
        global popup_body,popup_title,popup_Frame,xcord,back
        back=0
        popup_Frame=customtkinter.CTkFrame(master=app,corner_radius=10,bg_color='#000000', border_width=4,
                                           border_color='#37D0EE',fg_color='#414141',height=150,width=400)
        popup_Frame_child=customtkinter.CTkFrame(master=popup_Frame,corner_radius=10,bg_color='#414141', border_width=0,
                                           fg_color='#141414',height=100,width=380)
        popup_Frame_child.place(x=10,y=40)
        popup_Frame.grid_columnconfigure((0), weight=1)
        popup_Frame.place(x=150,y=200)
        popup_title=customtkinter.CTkLabel(popup_Frame,text=popup_title_text,
                                text_color='white',font=('Verdana',16))
        
        popup_body=customtkinter.CTkLabel(popup_Frame_child,text=popup_body_text,
                                text_color='grey',font=('Verdana',14))
        popup_body.place(x=15,y=15)
        popup_title.place(x=15,y=7)
        okimg=ImageTk.PhotoImage(Image.open("assets\\ok.png").resize((30,30)))
        noimg=ImageTk.PhotoImage(Image.open("assets\\no.png").resize((30,30)))
        #xcord=610

        #popup_Frame=customtkinter.CTkFrame(master=app,corner_radius=10,bg_color='#000000', border_width=2,
        #                                   fg_color='#37D0EE',height=60,width=340)
        #popup_Frame.place(x=xcord,y=100)
        #popup_anim()

        popup_title=customtkinter.CTkLabel(popup_Frame,text=popup_title_text,text_color='white',font=('Verdana',16))
        
        popup_body=customtkinter.CTkLabel(popup_Frame,text=popup_body_text,
                                text_color='grey',font=('Verdana',14))
        #popup_body.place(x=10,y=10)
        #popup_title.place(x=15,y=7)
        okimg=ImageTk.PhotoImage(Image.open("assets\\ok1w.png").resize((30,30)))
        noimg=ImageTk.PhotoImage(Image.open("assets\\no1w.png").resize((30,30)))
        def ok_button_event():

            disable_options(True)
            try:
                submit_button.configure(state="normal")
            except:
                pass
            try:
                upload_button.configure(state="normal")
            except:
                pass




            try:
                if idelete_button.cget('width')!=0:
                    print('DELETE ONLY')
                    delete_data()
                    popup_Frame.destroy()
                    intro_bool_dummy.toggle()
                    global noti_body_text
                    noti_body_text='Last  data  Entry  has \nbeen deleted successfully.'
                    noti_dummy.toggle()
            except:
                pass
            try:#upload confirmed
                

                
                
                if final_upload_button.cget('width')!=0:
                    print('UPDATE ONLY')
                    Crnt_Readings=[]
                    dummy=[]
                    dummy.append(entry5u.get())#
                    #Crnt_Readings.append(None)        #locked
                    dummy.append(entry4u.get())#
                    
                    dummy.append(entry1u.get())#locked
                    dummy.append(entry2u.get())#locked
                    dummy.append(entry3u.get())#locked
                    dummy.append(entry6u.get())
                    fields=["Meter no 5","Meter no 4","Meter no 1","Meter no 2","Meter no 3","Meter no 6","Current Reading","Total bill amount"]
                    dummy.append(entry7u.get())
                    dummy.append(entry8u.get())
                    
                    c=-1
                    
                    print(dummy)
                    for i in dummy:
                        c+=1
                        if i == '':
                            print('Compulsary field not entered !')
                            
                            noti_body_text="Please make sure a value has      \nbeen entered into '{}' field.".format(fields[c])
                            noti_dummy.toggle()
                            try:
                                popup_Frame.destroy()
                            except:
                                pass
                            print(100/0)#causing error 
                    c=-1
                    for i in dummy:
                        c+=1
                        try:
                            int(i)
                        except:
                            
                        
                            print('wrong tpye entered !')
                            noti_body_text="Only Integers are allowed        \nin '{}' field.".format(fields[c])
                            noti_dummy.toggle()
                            try:
                                popup_Frame.destroy()
                            except:
                                pass
                            print(100/0)#causing error 
                    dummy.pop()
                    dummy.pop()
                    for i in dummy:
                        Crnt_Readings.append(int(i))
                    print(Crnt_Readings)
                    #make all entrys int and noti for it    
                    Crnt_Readings.insert(1,None)        
                    if unitpriceu_box.get():
                        k=unitpriceu_field.get()
                        print("unitprice entry :",k,type(k))
                        if k!='':
                            a,b=get_lst_xl_vals()
                            update_xlsheet(a,Crnt_Readings,b,int(entry7u.get()),float(entry8u.get()),unitprice=float(k))

                            
                        else:
                            print('unit ticked but not entrerd ')
                            noti_body_text="Please Enter data in Unit Price field\nor untick  'Change Unitprice'."
                            noti_dummy.toggle()
                            try:
                                popup_Frame.destroy()
                            except:
                                pass
                            print(100/0)#causing error 
                    else:
                        
                        a,b=get_lst_xl_vals()
                        update_xlsheet(lst_readings=a,crnt_readings=Crnt_Readings,lst_NagarNigamReading=b,
                                        crnt_NagarNigamReading=int(entry7u.get()),NagarNigamBillAmt=float(entry8u.get()))
                    intro_bool_dummy.toggle()
                    noti_body_text='     All data has been updated\n        Successfully.'
                    noti_dummy.toggle()
                    
                
                
                
                '''
                ############################################
                if final_upload_button.cget('width')!=0:
                    Crntu_Readings=[]
                    print('UPLOAD ONLY')
                    Crntu_Readings.append(int(entry6u.get()))#
                    Crntu_Readings.append(None)        #locked
                    Crntu_Readings.append(int(entry4u.get()))#
                    Crntu_Readings.append(int(entry5u.get()))# 
                    Crntu_Readings.append(int(entry1u.get()))#locked
                    Crntu_Readings.append(int(entry2u.get()))#locked
                    Crntu_Readings.append(int(entry3u.get()))#locked
                                            #
                    if unitpriceu_box.get():
                        k=float(unitpriceu_field.get())
                        if k!='':
                            a,b=get_lst_xl_vals()
                            update_xlsheet(a,Crntu_Readings,b,int(entry7u.get()),float(entry8u.get()),unitprice=k)
                    else:
                        a,b=get_lst_xl_vals()
                        update_xlsheet(lst_readings=a,crnt_readings=Crntu_Readings,lst_NagarNigamReading=b,
                                        crnt_NagarNigamReading=int(entry7u.get()),NagarNigamBillAmt=float(entry8u.get()))
                    intro_bool_dummy.toggle()
                    noti_body_text='     All data has been updated\n        Successfully.'
                    noti_dummy.toggle()'''
                

                
            except:
                pass
            try:#bill gen confirmed
                #bill_button.configure(state='disabled',fg_color="#141414")
                if submit_button.cget('width')!=0:

                    

                    print('BILL ONLY')
                    Crnt_Readings=[]
                    dummy=[]
                    dummy.append(entry5.get())#locked
                    #Crnt_Readings.append(None)        #locked
                    dummy.append(entry4.get())#locked
                    dummy.append(entry1.get())#locked
                    dummy.append(entry2.get())#locked
                    dummy.append(entry3.get())#locked
                    dummy.append(entry6.get())#locked
                    fields=["Meter no 5","Meter no 4","Meter no 1","Meter no 2","Meter no 3","Meter no 6","Current Reading","Total bill amount"]
                    dummy.append(entry7.get())
                    dummy.append(entry8.get())
                    
                    c=-1
                    
                    print(dummy)
                    for i in dummy:
                        c+=1
                        if i == '':
                            print('Compulsary field not entered !')
                            
                            noti_body_text="Please make sure a value has      \nbeen entered into '{}' field.".format(fields[c])
                            noti_dummy.toggle()
                            try:
                                popup_Frame.destroy()
                            except:
                                pass
                            print(100/0)#causing error 
                    c=-1
                    for i in dummy:
                        c+=1
                        try:
                            int(i)
                        except:
                            
                        
                            print('wrong tpye entered !')
                            noti_body_text="Only Integers are allowed        \nin '{}' field.".format(fields[c])
                            noti_dummy.toggle()
                            try:
                                popup_Frame.destroy()
                            except:
                                pass
                            print(100/0)#causing error 
                    dummy.pop()
                    dummy.pop()
                    for i in dummy:
                        Crnt_Readings.append(int(i))
                    print(Crnt_Readings)
                    #make all entrys int and noti for it    
                    Crnt_Readings.insert(1,None)        
                    if unitprice_box.get():
                        k=unitprice_field.get()
                        print("unitprice entry :",k,type(k))
                        if k!='':
                            a,b=get_lst_xl_vals()
                            update_xlsheet(a,Crnt_Readings,b,int(entry7.get()),float(entry8.get()),unitprice=float(k))
                            global unitprice_int
                            unitprice_int=float(unitprice_field.get())
                            
                        else:
                            print('unit ticked but not entrerd ')
                            noti_body_text="Please Enter data in Unit Price field\nor untick  'Change Unitprice'."
                            noti_dummy.toggle()
                            try:
                                popup_Frame.destroy()
                            except:
                                pass
                            print(100/0)#causing error 
                    else:
                        
                        a,b=get_lst_xl_vals()
                        update_xlsheet(lst_readings=a,crnt_readings=Crnt_Readings,lst_NagarNigamReading=b,
                                        crnt_NagarNigamReading=int(entry7.get()),NagarNigamBillAmt=float(entry8.get()))

                    print("Shit updated ")
                    #(units,amts,unitprice_int,waterbool=wbool,tenants=nten)
                    global wbool,nten

                    wbool=water_box.get()
                    if wbool and water_field.get()!='':
                        nten=int(water_field.get())
                    elif wbool and water_field.get()=='':
                        nten=6
                    
                    try:
                        Main_Frame.destroy()
                    except:
                        pass
                    whatsapp_Frame=customtkinter.CTkFrame(master=app,corner_radius=10 ,fg_color='#000000',width=470,height=480)
                    whatsapp_Frame.place(x=120,y=10)
                    hint1_img =ImageTk.PhotoImage(Image.open("assets\\hint1.png").resize((300,164)))
                    hint2_img= ImageTk.PhotoImage(Image.open("assets\\hint2.png").resize((300,164)))
                    hint1_logo=customtkinter.CTkButton(whatsapp_Frame,image=hint1_img,text='' ,border_width=0,fg_color='transparent')
                    hint1_logo.place(x=7,y=10)
                    hint1_logo.configure(state='disabled')
                    hint2_logo=customtkinter.CTkButton(whatsapp_Frame,image=hint2_img,text='', border_width=0,fg_color='transparent')
                    hint2_logo.place(x=150,y=184)
                    hint2_logo.configure(state='disabled')
                    whatsapp_print_button= customtkinter.CTkButton(whatsapp_Frame,image=Printimg,
                                                                text='Print', border_width=0,text_color='grey',
                                                                fg_color='#141414',corner_radius=15,border_spacing=1,
                                                                    command=print_event_inbill,
                                                                        height=68,width=80,  hover_color='#424242',
                                                                    compound='left',font=('Verdana',12))
                    whatsapp_print_button.place(x=345,y=380)       
                    hint1_Bill=customtkinter.CTkLabel(whatsapp_Frame,text='Open Whatsapp\nweb & click\non a contact,\nas shown in img1.\nNote-\nAll the bills will\n be sent to the\nselected contact.',text_color='grey',font=('Verdana',14))
                    hint3_Bill=customtkinter.CTkLabel(whatsapp_Frame,text='Be patient, it takes about 30s to type\nup all the bills. This process will\nbe automated in future iterations.',text_color='grey',font=('Verdana',14))
                    hint4_Bill=customtkinter.CTkLabel(whatsapp_Frame,text='Step 1',text_color='white',font=('Verdana',15))
                    hint5_Bill=customtkinter.CTkLabel(whatsapp_Frame,text='Step 2',text_color='white',font=('Verdana',15))
                    hint2_Bill=customtkinter.CTkLabel(whatsapp_Frame,text=' Click on Print\nbutton in UI\nand quickyly change\nto whatsapp and\nclick on text box\nas shown in img2. ',text_color='grey',font=('Verdana',14))
                    hint1_Bill.place(x=330,y=40)
                    hint2_Bill.place(x=10,y=230)
                    hint3_Bill.place(x=40,y=385)
                    hint5_Bill.place(x=55,y=200)
                    hint4_Bill.place(x=370,y=10)
                    bill_button.configure(state='disabled',fg_color="#141414")
                    try:
                        noti_Frame.destroy()
                    except:
                        pass
                    
                    noti_body_text="Data has been updated! Proceed\nwith generating bill on whatsapp."
                    noti_dummy.toggle()
                    popup_Frame.destroy()


            except Exception as e:
                #bill_button.configure(state='disabled',fg_color="#141414")
                #bill_button.configure(state='disabled', fg_color='transparent')
                #print_button.configure(state='normal', fg_color='transparent')
                #intro_button.configure(state='normal', fg_color='transparent')
                #upload_button.configure(state='normal', fg_color='transparent')
                #delete_button.configure(state='normal', fg_color='transparent')
                print('bill fail')
                

        def no_button_event():
            disable_options(True)
            try:
                submit_button.configure(state="disabled")
            except:
                pass
            try:
                final_upload_button.configure(state="disabled")
            except:
                pass
            popup_Frame.destroy()
            try:
                noti_Frame.destroy()
            except:
                pass
            try:
                submit_button.configure(state='normal')
            except:
                pass
            try:
                final_upload_button.configure(state='normal')
            except:
                pass

            '''        ok_button=customtkinter.CTkButton(popup_Frame,image= okimg,
                                        corner_radius=15,border_spacing=0,text='',hover_color='#414141',
                                        fg_color='transparent',command=ok_button_event,
                                        height=35,width=35,  border_width=0) 
        ok_button.place(x=100,y=10 )
        no_button=customtkinter.CTkButton(popup_Frame,image= noimg,
                                        corner_radius=15,border_spacing=0,text='',hover_color='#414141',
                                        fg_color='transparent',command=no_button_event,
                                        height=35,width=10,  border_width=0) 
        '''
        ok_button=customtkinter.CTkButton(popup_Frame,image= okimg,
                                        corner_radius=10,border_spacing=0,text=None,hover_color='green',
                                        fg_color='#141414',command=ok_button_event,border_color="#FFFFFF",
                                        height=45,width=10,  border_width=1,bg_color="#141414") 
        ok_button.place(x=332,y=44 )
        no_button=customtkinter.CTkButton(popup_Frame,image= noimg,
                                        corner_radius=10,border_spacing=0,text=None,hover_color='red',
                                        fg_color='#141414',command=no_button_event,border_color="#FFFFFF",
                                        height=45,width=10,  border_width=1,bg_color="#141414") 
        no_button.place(x=332,y=91 )       
    popup_bool_dummy=customtkinter.CTkCheckBox(Option_Frame,command=popup)

    

    
    #Intro Page
    
    global intro_bool_dummy
    def intro_page():
        try:
            noti_Frame.destroy()
        except:
            pass
        try:
            Main_Frame.cget("width")
            Main_Frame.destroy()
        except:
            pass
        try:
            update_Frame.destroy()
        except:
            pass
        try:
            Print_Frame.destroy()
        except:
            pass
        try:
            Delete_Frame.cget('width')
            Delete_Frame.destroy()
        except:
            pass
        intro_button.configure(state='disabled',fg_color="#141414")
        bill_button.configure(state='normal', fg_color='transparent')
        print_button.configure(state='normal', fg_color='transparent')
        delete_button.configure(state='normal', fg_color='transparent')
        upload_button.configure(state='normal', fg_color='transparent')
        global Intro_Frame
        Intro_Frame=customtkinter.CTkFrame(master=app,corner_radius=10,fg_color='#000000',width=470,height=480)
        Intro_Frame.place(x=120,y=10)
        logoimg=ImageTk.PhotoImage(Image.open("assets\\logo.png"))
        intro_logo=customtkinter.CTkButton(Intro_Frame,image=logoimg,text='', state='disabled',border_width=0,fg_color='transparent')
        intro_logo.place(x=70,y=20)


        hint1_intro=customtkinter.CTkLabel(Intro_Frame,text='Click "Bill" to upload data and generate a new bill. ',text_color='grey',font=('Verdana',14))
        hint2_intro=customtkinter.CTkLabel(Intro_Frame,text='Click "Print" to generate an older bill.',text_color='grey',font=('Verdana',14))
        hint3_intro=customtkinter.CTkLabel(Intro_Frame,text='Click "Update" to upload bill into database and not print.',text_color='grey',font=('Verdana',14))    
        
        hint1_intro.place(x=30,y=300)
        hint2_intro.place(x=30,y=350)
        hint3_intro.place(x=30,y=400)
        print(hint1_intro.cget("width"))
    intro_bool_dummy=customtkinter.CTkCheckBox(Option_Frame,command=intro_page)

    #Bill Page 
    def bill_page():

        try:
            Intro_Frame.cget('width')
            Intro_Frame.destroy()
        except:
            pass
        try:
            update_Frame.destroy()
        except:
            pass
        try:
            Print_Frame.destroy()
        except:
            pass
        try:
            Delete_Frame.cget('width')
            Delete_Frame.destroy()
        except:
            pass
        global water_box, check_Frame,water_box,unitprice_box, Main_Frame,btitle
        global entry1,entry2,entry3,entry4,entry5,entry6,entry7,entry8,submit_button,bill_logo
        Main_Frame=customtkinter.CTkFrame(master=app,corner_radius=10 ,fg_color='#000000',width=470,height=490)

        Main_Frame.place(x=120,y=10)                    
        entry_Frame=customtkinter.CTkFrame(master=Main_Frame, fg_color='#141414',corner_radius=10)    
        check_Frame=customtkinter.CTkFrame(master=Main_Frame, fg_color='#141414',corner_radius=10)        
        entry_Frame.grid_columnconfigure((0), weight=1)
        check_Frame.grid_columnconfigure((0), weight=1)
        entry_Frame.pack(padx=(10,5),pady=(120,10) , side='left',fill='both')
        check_Frame.pack(padx=(5,10),pady=(120,130) )

        submitimg=ImageTk.PhotoImage(Image.open("assets\\submit.png").resize((40,40)))
        submit_button=customtkinter.CTkButton(Main_Frame,image= submitimg,
                                        corner_radius=15,border_spacing=1,text='Submit',
                                        fg_color='#008AA9',command=submit_button_event,
                                        height=68,width=100,  border_width=0,
                                        compound='left',font=('Verdana',14)) 
        submit_button.place(x=300,y=380 )
        
        Billimgb =ImageTk.PhotoImage(Image.open("assets\\bill.png").resize((64,64)))
        bill_logo=customtkinter.CTkButton(Main_Frame,text='',state='disabled',image=Billimgb, border_width=0,fg_color='transparent')
        bill_logo.place(x=50,y=20)
        btitle = customtkinter.CTkLabel(Main_Frame,text='Upload Data and Generate \nBill on Whatsapp.',text_color='grey',font=('Verdana',18))
        btitle.place(x=150,y=35)

        btitle_entry=customtkinter.CTkLabel(entry_Frame,text='Enter current readings for :',text_color='grey',font=('Verdana',14))
        btitle_entry.grid(row=0, column=1, padx=20, pady=(15,5), sticky='nw')

        entry1=customtkinter.CTkEntry(entry_Frame, placeholder_text="Meter No. 1",placeholder_text_color='grey',height=15,width=150,corner_radius=5)
        entry1.grid(row=1, column=1, padx=20, pady=(0,5))

        entry2=customtkinter.CTkEntry(entry_Frame, placeholder_text="Meter No. 2",placeholder_text_color='grey',height=15,width=150,corner_radius=5)
        entry2.grid(row=2, column=1, padx=20, pady=5)

        entry3=customtkinter.CTkEntry(entry_Frame, placeholder_text="Meter No. 3",placeholder_text_color='grey',height=15,width=150,corner_radius=5)
        entry3.grid(row=3, column=1, padx=20, pady=5)

        entry4=customtkinter.CTkEntry(entry_Frame, placeholder_text="Meter No. 4",placeholder_text_color='grey',height=15,width=150,corner_radius=5)
        entry4.grid(row=4, column=1, padx=20, pady=5)

        entry5=customtkinter.CTkEntry(entry_Frame, placeholder_text="Meter No. 5",placeholder_text_color='grey',height=15,width=150,corner_radius=5)
        entry5.grid(row=5, column=1, padx=20, pady=5)

        entry6=customtkinter.CTkEntry(entry_Frame, placeholder_text="Meter No. 0",placeholder_text_color='grey',height=15,width=150,corner_radius=5)
        entry6.grid(row=6, column=1, padx=20, pady=5)

        title_entry1=customtkinter.CTkLabel(entry_Frame,text='Values form NagarNigam bill :',text_color='grey',font=('Verdana',14))
        title_entry1.grid(row=7, column=1, padx=20, pady=(15,5))
        
        entry7=customtkinter.CTkEntry(entry_Frame, placeholder_text="Current Reading",placeholder_text_color='grey',height=15,width=150,corner_radius=5)
        entry7.grid(row=8, column=1, padx=20, pady=5)

        entry8=customtkinter.CTkEntry(entry_Frame, placeholder_text="Total Bill Amount",placeholder_text_color='grey',height=15,width=150,corner_radius=5)
        entry8.grid(row=9, column=1, padx=20, pady=(5,10))
        

        water_box=customtkinter.CTkCheckBox(check_Frame,height=2,fg_color='#37D0EE',width=2,text='Charge for water',text_color='grey',font=('Verdana',14),command=water_need_fn)
        water_box.grid(row=1,column=1,padx=20,pady=10)
        water_box.toggle()
        unitprice_box=customtkinter.CTkCheckBox(check_Frame,height=2,fg_color='#37D0EE',width=2,text='Change Unit Price',text_color='grey',font=('Verdana',14),command=unitprice_box_fn)
        unitprice_box.grid(row=4,column=1,padx=20,pady=10)
        unitprice_box.toggle()
    bill_bool_dummy=customtkinter.CTkCheckBox(Option_Frame,command=bill_page)
    update_dummy=customtkinter.CTkCheckBox(Option_Frame,command=update_page)  
    
    #Sidebar frame
    Billimg=ImageTk.PhotoImage(Image.open("assets\\bill1.png").resize((40,40)))
    Printimg=ImageTk.PhotoImage(Image.open("assets\\print.png").resize((40,40)))
    Deleteimg=ImageTk.PhotoImage(Image.open("assets\\delete.png").resize((40,40)))
    Optionsimg=ImageTk.PhotoImage(Image.open("assets\\options1.png").resize((40,40)))
    Uploadimg=ImageTk.PhotoImage(Image.open("assets\\upload.png").resize((40,40)))
 

    
    global intro_button,bill_button,print_button,delete_button,upload_button
    intro_button=customtkinter.CTkButton(Option_Frame,image= Optionsimg,
                                      corner_radius=15,border_spacing=0,text='',
                                      fg_color='transparent',command=intro_page,
                                      height=70,width=70,  hover_color='#141414',
                                      compound='left',font=('Verdana',12))
    intro_button.place(x=14,y=22)

    bill_button=customtkinter.CTkButton(Option_Frame,image= Billimg,
                                      corner_radius=15,border_spacing=1,text='',
                                      fg_color='transparent',command=bill_button_event,
                                      height=70,width=70,  hover_color='#141414',
                                      compound='left',font=('Verdana',12))  
    bill_button.place(x=14,y=114)
    
    print_button=customtkinter.CTkButton(Option_Frame,image= Printimg,
                                      corner_radius=15,border_spacing=1,text='',
                                      fg_color='transparent',command=prnt_button_event,
                                      height=70,width=70,  hover_color='#141414',
                                      compound='left',font=('Verdana',12))   
    print_button.place(x=14,y=298)   

    delete_button=customtkinter.CTkButton(Option_Frame,image= Deleteimg,
                                      corner_radius=15,border_spacing=1,text='',
                                      fg_color='transparent',command=del_button_event,
                                      height=70,width=70,  hover_color='#141414',
                                      compound='left',font=('Verdana',12))    
    delete_button.place(x=14,y=390)

    upload_button=customtkinter.CTkButton(Option_Frame,image= Uploadimg,
                                      corner_radius=15,border_spacing=1,text='',
                                      fg_color='transparent',command=uupload_button_event,
                                      height=70,width=70,  hover_color='#141414',
                                      compound='left',font=('Verdana',12))    
    upload_button.place(x=14,y=206)    
    intro_bool_dummy.toggle()

    global noti_body_text,noti_img_bool
    img=ImageTk.PhotoImage(Image.open("assets\\pray.png").resize((45,45)))

    noti_body_text='           Welcome!!! to VHEB'
    
    noti_dummy.toggle()
    l=customtkinter.CTkButton(noti_Frame,image= img,
                                        corner_radius=7,border_spacing=0,text=None,
                                        fg_color='transparent',state='disabled',
                                         height=45,width=10, border_width=0)
    l.place(x=3,y=3)

    global drive_body_text
    drive_body_text='G Drive Connection has been\nestablished successfully'
    drive_dummy.toggle()

    app.mainloop()
file_download()
for i in range(10):
    print('Start!!!')
__main__()
file_upload()

for i in range(10):
    print('END!!!')